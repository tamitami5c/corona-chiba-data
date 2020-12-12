import urllib
import os
import json
import csv
import datetime

import requests
from openpyxl import load_workbook,utils

def get_file_name(url):
    path=urllib.parse.urlparse(url).path
    return os.path.split(path)[-1]

def download_file(url):
    file_name=get_file_name(url)

    response=requests.get(url)
    if response.status_code!=requests.codes.ok:
        raise Exception("status_code!=200")
    response.encoding=response.apparent_encoding

    with open(file_name,"wb") as f:
        f.write(response.content)
    return file_name

def daterange(start_date,end_date):
    """
    start_date(含む)からend_date(含む)まで
    """
    for n in range(int((end_date-start_date).days)+1):
        yield start_date+datetime.timedelta(n)

def main():
    #千葉県のファイル
    url="https://www.pref.chiba.lg.jp/shippei/press/2019/documents/chiba_corona_data.xlsx"
    file_name=download_file(url)

    wb=load_workbook(file_name,data_only=True)
    ws=wb["新型コロナウイルス感染者数（検査確定日、公表日、7日間平均）"]
    d=[]
    max_row=ws.max_row
    for row in range(5,max_row+1):
        #公表日
        dt=ws.cell(row,5).value
        if dt is None:
            break

        #時間にintのものが混じっていたのでdatetimeに変換
        if isinstance(dt,int):
            dt=utils.datetime.from_excel(dt)
        date=dt.date()
        #感染者数
        count=ws.cell(row,6).value
        print(date,",",count)

        d.append({
            "date":date.isoformat(),
            "count":count,
        })


    #jsonに書きだし
    data={
        "data":d
    }
    print(data)

    with open("chiba_data.json","w") as f:
        json.dump(data,f,indent=4)

if __name__ == "__main__":
    main()